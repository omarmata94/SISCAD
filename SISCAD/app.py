import math
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook

WORKBOOK_FILE = Path("ListaDeCalificaciones3E.xlsx")
TRIMESTER_SHEETS = ["1er Trimestre", "2do Trimestre", "3er Trimestre"]
ACT_COLUMNS = [f"Actividad {i}" for i in range(1, 31)]
SECRETARIA_ACTIVITY_COUNTS = {
    "1er Trimestre": 11,
    "2do Trimestre": 20,
    "3er Trimestre": 25,
}
SECRETARIA_TRIMESTER_SHEET_NAMES = {
    "1er Trimestre": "Primer_Trimestre",
    "2do Trimestre": "Segundo_Trimestre",
    "3er Trimestre": "Tercer_Trimestre",
}
SECRETARIA_TIRILLA_NAMES = {
    "1er Trimestre": "TIRILLA T1",
    "2do Trimestre": "TIRILLA T2",
    "3er Trimestre": "TIRILLA T3",
}
SECRETARIA_SUBJECT = "TECNOLOGIA"
TASK_COLUMNS = [f"Tarea {i}" for i in range(1, 31)]
PROJECT_COLUMNS = [f"Proyecto {i}" for i in range(1, 8)]
METRIC_COLUMNS = ["Participación", "Conducta", "Evaluación (Examen)"]
WEIGHTS = {
    "Promedio Actividades": 0.2,
    "Promedio Tareas": 0.2,
    "Promedio Proyectos": 0.2,
    "Participación": 0.1,
    "Conducta": 0.1,
    "Evaluación (Examen)": 0.2,
}


def get_editor():
    if hasattr(st, "data_editor"):
        return st.data_editor
    if hasattr(st, "experimental_data_editor"):
        return st.experimental_data_editor
    return None


def parse_datos_sheet(wb):
    ws = wb["DATOS"]
    general = {
        "Escuela": ws.cell(row=3, column=4).value,
        "Docente": ws.cell(row=4, column=4).value,
        "Grado": ws.cell(row=5, column=4).value,
        "Grupo": ws.cell(row=6, column=4).value,
    }
    students = []
    for row in ws.iter_rows(min_row=11, max_row=100, min_col=2, max_col=3, values_only=True):
        numero, nombre = row
        if numero is None and nombre is None:
            continue
        if nombre is None:
            break
        students.append({
            "N°": int(numero) if isinstance(numero, (int, float)) and not math.isnan(numero) else None,
            "Nombre": nombre,
        })
    return general, pd.DataFrame(students)


def find_trimester_columns(ws):
    header_labels = list(ws.iter_rows(min_row=9, max_row=9, max_col=100, values_only=True))[0]
    header_values = list(ws.iter_rows(min_row=10, max_row=10, max_col=100, values_only=True))[0]

    section = None
    activity_cols = []
    task_cols = []
    project_cols = []
    promedio_actividades_col = None
    promedio_tareas_col = None
    promedio_proyectos_col = None
    metric_positions = {}
    final_col = None

    for col_idx, (label, value) in enumerate(zip(header_labels, header_values), start=1):
        label_text = str(label).strip().upper() if isinstance(label, str) else ""
        value_text = str(value).strip().upper() if isinstance(value, str) else ""

        if label_text == "ACTIVIDADES":
            section = "activities"
        if label_text == "TAREAS":
            section = "tasks"
        if label_text == "PROYECTOS":
            section = "projects"
        if "PROMEDIO ACTIVIDADES" in label_text or "PROMEDIO ACTIVIDADES" in value_text:
            promedio_actividades_col = col_idx
            section = "tasks"
            continue
        if "PROMEDIO TAREAS" in label_text or "PROMEDIO TAREAS" in value_text:
            promedio_tareas_col = col_idx
            section = "projects"
            continue
        if "PROMEDIO PROYECTOS" in label_text or "PROMEDIO PROYECTOS" in value_text:
            promedio_proyectos_col = col_idx
            section = "metrics"
            continue
        if "PARTICIPACI" in label_text:
            metric_positions["Participación"] = col_idx
            section = "metrics"
            continue
        if "CONDUCTA" in label_text:
            metric_positions["Conducta"] = col_idx
            section = "metrics"
            continue
        if "EVALUACIÓN" in label_text:
            metric_positions["Evaluación (Examen)"] = col_idx
            section = "metrics"
            continue
        if "CALIF FINAL" in label_text:
            final_col = col_idx
            continue

        if isinstance(value, (int, float)):
            if section == "activities":
                activity_cols.append(col_idx)
            elif section == "tasks":
                task_cols.append(col_idx)
            elif section == "projects":
                project_cols.append(col_idx)

    activity_cols = activity_cols[:30]
    task_cols = task_cols[:30]
    project_cols = project_cols[:7]

    if final_col is None:
        for col_idx, label in enumerate(header_labels, start=1):
            if isinstance(label, str) and "CALIF FINAL" in label.upper():
                final_col = col_idx
                break

    return {
        "activity_cols": activity_cols,
        "task_cols": task_cols,
        "project_cols": project_cols,
        "promedio_actividades_col": promedio_actividades_col,
        "promedio_tareas_col": promedio_tareas_col,
        "promedio_proyectos_col": promedio_proyectos_col,
        "metric_positions": metric_positions,
        "final_col": final_col,
    }


def parse_trimester_sheet(wb, sheet_name):
    ws = wb[sheet_name]
    columns = find_trimester_columns(ws)
    rows = list(ws.iter_rows(min_row=11, max_row=100, max_col=100, values_only=True))
    data = []

    for row in rows:
        if row[0] is None and row[1] is None:
            continue
        if row[1] is None:
            continue
        record = {
            "N°": int(row[0]) if isinstance(row[0], (int, float)) and not math.isnan(row[0]) else None,
            "Nombre": row[1],
        }

        for idx, col_idx in enumerate(columns["activity_cols"], start=1):
            record[f"Actividad {idx}"] = row[col_idx - 1]
        for idx in range(len(columns["activity_cols"]) + 1, 31):
            record[f"Actividad {idx}"] = None

        record["Promedio Actividades"] = row[columns["promedio_actividades_col"] - 1] if columns["promedio_actividades_col"] else None

        for idx, col_idx in enumerate(columns["task_cols"], start=1):
            record[f"Tarea {idx}"] = row[col_idx - 1]
        for idx in range(len(columns["task_cols"]) + 1, 31):
            record[f"Tarea {idx}"] = None

        record["Promedio Tareas"] = row[columns["promedio_tareas_col"] - 1] if columns["promedio_tareas_col"] else None

        for idx, col_idx in enumerate(columns["project_cols"], start=1):
            record[f"Proyecto {idx}"] = row[col_idx - 1]
        for idx in range(len(columns["project_cols"]) + 1, 6):
            record[f"Proyecto {idx}"] = None

        record["Promedio Proyectos"] = row[columns["promedio_proyectos_col"] - 1] if columns["promedio_proyectos_col"] else None

        for metric_name in METRIC_COLUMNS:
            metric_col = columns["metric_positions"].get(metric_name)
            record[metric_name] = row[metric_col - 1] if metric_col else None

        record["Calif Final"] = row[columns["final_col"] - 1] if columns["final_col"] else None
        data.append(record)

    all_columns = ["N°", "Nombre"] + ACT_COLUMNS + ["Promedio Actividades"] + TASK_COLUMNS + ["Promedio Tareas"] + PROJECT_COLUMNS + ["Promedio Proyectos"] + METRIC_COLUMNS + ["Calif Final"]
    return pd.DataFrame(data, columns=all_columns)


def normalize_grades(df):
    result = df.copy()
    for col in ACT_COLUMNS + TASK_COLUMNS + PROJECT_COLUMNS + METRIC_COLUMNS:
        if col in result.columns:
            result[col] = pd.to_numeric(result[col], errors="coerce")
    return result


def compute_trimester_result(df):
    df = normalize_grades(df)
    df["Promedio Actividades"] = df[ACT_COLUMNS].mean(axis=1, skipna=True)
    df["Promedio Tareas"] = df[TASK_COLUMNS].mean(axis=1, skipna=True)
    df["Promedio Proyectos"] = df[PROJECT_COLUMNS].mean(axis=1, skipna=True)
    for metric in METRIC_COLUMNS:
        df[metric] = pd.to_numeric(df[metric], errors="coerce")
    df["Calif Final"] = (
        df["Promedio Actividades"] * WEIGHTS["Promedio Actividades"]
        + df["Promedio Tareas"] * WEIGHTS["Promedio Tareas"]
        + df["Promedio Proyectos"] * WEIGHTS["Promedio Proyectos"]
        + df["Participación"] * WEIGHTS["Participación"]
        + df["Conducta"] * WEIGHTS["Conducta"]
        + df["Evaluación (Examen)"] * WEIGHTS["Evaluación (Examen)"]
    )
    return df


def build_concentrado_sheet(trimester_dfs):
    if not trimester_dfs:
        return pd.DataFrame()
    merged = None
    for sheet_name, df in trimester_dfs.items():
        partial = df[["N°", "Nombre", "Calif Final"]].copy()
        partial = partial.rename(columns={"Calif Final": sheet_name})
        if merged is None:
            merged = partial
        else:
            merged = pd.merge(merged, partial, on=["N°", "Nombre"], how="outer")
    merged["Promedio General"] = merged[[sheet for sheet in trimester_dfs.keys()]].mean(axis=1, skipna=True)
    return merged


def build_resultados_sheet(general, concentrado_df):
    summary = {
        "Escuela": [general.get("Escuela")],
        "Docente": [general.get("Docente")],
        "Grado": [general.get("Grado")],
        "Grupo": [general.get("Grupo")],
    }
    for sheet_name in TRIMESTER_SHEETS:
        if sheet_name in concentrado_df.columns:
            summary[f"Promedio Grupal {sheet_name}"] = [concentrado_df[sheet_name].mean(skipna=True)]
        else:
            summary[f"Promedio Grupal {sheet_name}"] = [None]
    summary["Promedio General Grupal"] = [concentrado_df["Promedio General"].mean(skipna=True) if "Promedio General" in concentrado_df.columns else None]
    return pd.DataFrame(summary)


def get_template_workbook(source):
    if source is None:
        return None
    if isinstance(source, (bytes, bytearray)):
        return load_workbook(BytesIO(source), data_only=False)
    return load_workbook(source, data_only=False)


def update_datos_sheet(ws, general, students_df):
    ws.cell(row=3, column=4, value=general.get("Escuela"))
    ws.cell(row=4, column=4, value=general.get("Docente"))
    ws.cell(row=5, column=4, value=general.get("Grado"))
    ws.cell(row=6, column=4, value=general.get("Grupo"))

    start_row = 11
    for idx, student in students_df.iterrows():
        ws.cell(row=start_row + idx, column=2, value=student.get("N°"))
        ws.cell(row=start_row + idx, column=3, value=student.get("Nombre"))

    for row_idx in range(start_row + len(students_df), ws.max_row + 1):
        ws.cell(row=row_idx, column=2, value=None)
        ws.cell(row=row_idx, column=3, value=None)


def update_trimester_sheet(ws, df):
    columns = find_trimester_columns(ws)
    start_row = 11

    for idx, row in df.iterrows():
        excel_row = start_row + idx
        ws.cell(row=excel_row, column=1, value=row.get("N°"))

        name_cell = ws.cell(row=excel_row, column=2)
        if name_cell.data_type != "f":
            name_cell.value = row.get("Nombre")

        for act_idx, col_idx in enumerate(columns["activity_cols"], start=1):
            ws.cell(row=excel_row, column=col_idx, value=row.get(f"Actividad {act_idx}"))

        for task_idx, col_idx in enumerate(columns["task_cols"], start=1):
            ws.cell(row=excel_row, column=col_idx, value=row.get(f"Tarea {task_idx}"))

        for proj_idx, col_idx in enumerate(columns["project_cols"], start=1):
            ws.cell(row=excel_row, column=col_idx, value=row.get(f"Proyecto {proj_idx}"))

        for metric_name, col_idx in columns["metric_positions"].items():
            ws.cell(row=excel_row, column=col_idx, value=row.get(metric_name))

        if columns["final_col"] is not None:
            final_cell = ws.cell(row=excel_row, column=columns["final_col"])
            if final_cell.data_type != "f":
                final_cell.value = row.get("Calif Final")

    for row_idx in range(start_row + len(df), ws.max_row + 1):
        ws.cell(row=row_idx, column=1, value=None)
        name_cell = ws.cell(row=row_idx, column=2)
        if name_cell.data_type != "f":
            name_cell.value = None

        for col_idx in columns["activity_cols"] + columns["task_cols"] + columns["project_cols"]:
            ws.cell(row=row_idx, column=col_idx, value=None)

        for col_idx in columns["metric_positions"].values():
            ws.cell(row=row_idx, column=col_idx, value=None)


def save_workbook(general, students_df, trimester_dfs, concentrado_df, resultados_df, output_path, template_source=None):
    template_wb = get_template_workbook(template_source)
    if template_wb is not None:
        if "DATOS" in template_wb.sheetnames:
            update_datos_sheet(template_wb["DATOS"], general, students_df)
        for sheet_name, df in trimester_dfs.items():
            if sheet_name in template_wb.sheetnames:
                update_trimester_sheet(template_wb[sheet_name], df)
        template_wb.save(output_path)
        return

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        datos_rows = [
            [None, "DATOS GENERALES", None, None],
            [None, "NOMBRE DE LA ESCUELA:", None, general.get("Escuela")],
            [None, "NOMBRE DEL DOCENTE:", None, general.get("Docente")],
            [None, "GRADO:", None, general.get("Grado")],
            [None, "GRUPO:", None, general.get("Grupo")],
            [None, None, None, None],
            [None, "ALUMNOS", None, None],
            [None, "N°", "Nombre", None],
        ]
        pd.DataFrame(datos_rows).to_excel(writer, sheet_name="DATOS", header=False, index=False)
        students_df.to_excel(writer, sheet_name="DATOS", header=False, index=False, startrow=len(datos_rows))
        for sheet_name, df in trimester_dfs.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        concentrado_df.to_excel(writer, sheet_name="CONCENTRADO", index=False)
        resultados_df.to_excel(writer, sheet_name="RESULTADOS", index=False)


def format_secretaria_grade(value):
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def build_secretaria_group_title(general):
    grado_value = general.get("Grado", "")
    if isinstance(grado_value, (int, float)) and not math.isnan(grado_value):
        grado = str(int(grado_value)) if float(grado_value).is_integer() else str(grado_value)
    else:
        grado = str(grado_value).strip()

    grupo = str(general.get("Grupo", "")).strip()
    if not grado:
        return None
    if not grupo:
        return f"{grado}°"
    if len(grupo) == 1:
        return f'{grado}° "   {grupo}   "'
    return f'{grado}° "{grupo}"'


def round_secretaria_grade(value):
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    return int(value + 0.5)


def build_secretaria_trimester_rows(trimester_name, trimester_df, general):
    activity_count = SECRETARIA_ACTIVITY_COUNTS.get(trimester_name, 11)
    activity_columns = ACT_COLUMNS[:activity_count]

    header_row = [
        "No.",
        None,
        "Alumno",
        "A/B",
        "DIAG.",
    ] + [float(i) for i in range(1, activity_count + 1)] + [
        None,
        "ACT.",
        "PART.",
        "LEM",
        "PROY.",
        "EX.",
        "SUMA",
        "PROM.",
        None,
        "No.",
        "FALTAS",
        "PROM.",
    ]

    rows = []
    for _, row in trimester_df.iterrows():
        activity_values = [format_secretaria_grade(row[col]) for col in activity_columns]
        valid_activities = [value for value in activity_values if value is not None]
        promedio_actividades = None
        if valid_activities:
            promedio_actividades = sum(valid_activities) / len(valid_activities)

        participacion = format_secretaria_grade(row.get("Participación"))
        conducta = format_secretaria_grade(row.get("Conducta"))
        promedio_proyectos = format_secretaria_grade(row.get("Promedio Proyectos"))
        examen = format_secretaria_grade(row.get("Evaluación (Examen)"))

        suma = None
        prom = None
        if None not in (promedio_actividades, participacion, conducta, promedio_proyectos, examen):
            suma = promedio_actividades + participacion + conducta + promedio_proyectos + examen
            prom = suma / 5

        final_grade = round_secretaria_grade(prom)

        rows.append([
            format_secretaria_grade(row.get("N°")),
            None,
            row.get("Nombre"),
            None,
            None,
            *activity_values,
            None,
            promedio_actividades,
            participacion,
            conducta,
            promedio_proyectos,
            examen,
            suma,
            prom,
            None,
            format_secretaria_grade(row.get("N°")),
            0,
            final_grade,
        ])

    return header_row, rows


def build_secretaria_final_rows(trimester_dfs, general):
    final_sheet_name = "FINAL"
    if any(sheet_name not in trimester_dfs for sheet_name in TRIMESTER_SHEETS):
        return None, []

    final_df = build_concentrado_sheet({
        sheet_name: trimester_dfs[sheet_name][["N°", "Nombre", "Calif Final"]].copy()
        for sheet_name in TRIMESTER_SHEETS
    })
    final_df["1er Trimestre"] = final_df["1er Trimestre"].apply(round_secretaria_grade)
    final_df["2do Trimestre"] = final_df["2do Trimestre"].apply(round_secretaria_grade)
    final_df["3er Trimestre"] = final_df["3er Trimestre"].apply(round_secretaria_grade)
    final_df["Promedio Final"] = final_df[TRIMESTER_SHEETS].mean(axis=1, skipna=True)
    final_df = final_df.sort_values("N°", na_position="last")

    counts = {
        "menor o igual a 6.9": int(((final_df["Promedio Final"] <= 6.9) & final_df["Promedio Final"].notna()).sum()),
        "7 a 7.9": int(((final_df["Promedio Final"] >= 7.0) & (final_df["Promedio Final"] <= 7.9)).sum()),
        "8 a 10": int((final_df["Promedio Final"] >= 8.0).sum()),
    }

    header_row = [
        "No.",
        None,
        "Alumno",
        "A/B",
        None,
        "FALTAS",
        "PROM.",
        None,
        "FALTAS",
        "PROM.",
        None,
        "FALTAS",
        "PROM.",
        None,
        "FALTAS",
        "PROM.",
        None,
        "menor o igual a 6.9",
        counts["menor o igual a 6.9"],
        None,
    ]

    rows = []
    for index, (_, row) in enumerate(final_df.iterrows()):
        prom1 = round_secretaria_grade(row.get("1er Trimestre"))
        prom2 = round_secretaria_grade(row.get("2do Trimestre"))
        prom3 = round_secretaria_grade(row.get("3er Trimestre"))
        prom_final = format_secretaria_grade(row.get("Promedio Final"))

        classification_label = None
        classification_count = None
        if index == 0:
            classification_label = "7 a 7.9"
            classification_count = counts["7 a 7.9"]
        elif index == 1:
            classification_label = "8 a 10"
            classification_count = counts["8 a 10"]

        rows.append([
            format_secretaria_grade(row.get("N°")),
            None,
            row.get("Nombre"),
            None,
            None,
            0,
            prom1,
            None,
            0,
            prom2,
            None,
            0,
            prom3,
            None,
            0,
            prom_final,
            None,
            classification_label,
            classification_count,
            None,
        ])

    return header_row, rows


def build_secretaria_tirilla_rows(trimester_name, trimester_df, general):
    header_row = [None, "No.", "NOMBRE DEL ALUMNO", "PROMEDIO"]
    rows = []
    for _, row in trimester_df.iterrows():
        prom = round_secretaria_grade(row.get("Calif Final"))
        rows.append([None, format_secretaria_grade(row.get("N°")), row.get("Nombre"), prom])
    return header_row, rows


def save_secretaria_workbook(selected_trimesters, general, trimester_dfs, output_path):
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    for trimester_name in selected_trimesters:
        if trimester_name not in SECRETARIA_TRIMESTER_SHEET_NAMES:
            continue
        sheet_name = SECRETARIA_TRIMESTER_SHEET_NAMES[trimester_name]
        ws = wb.create_sheet(title=sheet_name)
        ws.append([None, None, f"Asignatura: {SECRETARIA_SUBJECT}"])
        ws.append([None, None, f"Docente: {general.get('Docente', '')}"])
        ws.append([None, None, None])
        group_title = build_secretaria_group_title(general)
        ws.append([None, "GÉNERO", group_title, None, None, "Calificación Actividades"])
        header_row, rows = build_secretaria_trimester_rows(trimester_name, trimester_dfs[trimester_name], general)
        ws.append(header_row)
        for row in rows:
            ws.append(row)

        tirilla_name = SECRETARIA_TIRILLA_NAMES[trimester_name]
        ws2 = wb.create_sheet(title=tirilla_name)
        ws2.append([None, None, f"ESCUELA: {general.get('Escuela', '')}"])
        ws2.append([None, f"Docente: {general.get('Docente', '')}"])
        ws2.append([None, f"TRIM. {trimester_name[0]}", None, general.get("Grupo", "")])
        ws2.append([None, None, None])
        header_row, rows = build_secretaria_tirilla_rows(trimester_name, trimester_dfs[trimester_name], general)
        ws2.append(header_row)
        for row in rows:
            ws2.append(row)

    if set(SECRETARIA_TRIMESTER_SHEET_NAMES.keys()).issubset(set(selected_trimesters)):
        ws_final = wb.create_sheet(title="FINAL")
        ws_final.append([None, None, f"Asignatura: {SECRETARIA_SUBJECT}"])
        ws_final.append([None, None, f"Docente: {general.get('Docente', '')}"])
        ws_final.append([None, None, None])
        group_title = build_secretaria_group_title(general)
        ws_final.append([
            None,
            None,
            group_title,
            None,
            None,
            "PRIMER TRIMESTRE",
            None,
            None,
            "SEGUNDO TRIMESTRE",
            None,
            None,
            "TERCER TRIMESTRE",
            None,
            None,
            "FINAL",
            None,
            None,
            None,
            None,
            None,
        ])
        header_row, rows = build_secretaria_final_rows(trimester_dfs, general)
        if header_row is not None:
            ws_final.append(header_row)
            for row in rows:
                ws_final.append(row)

    wb.save(output_path)


def load_workbook_data(source):
    wb = load_workbook(source, data_only=True)
    general, students_df = parse_datos_sheet(wb)
    trimester_data = {sheet: parse_trimester_sheet(wb, sheet) for sheet in TRIMESTER_SHEETS}
    return general, students_df, trimester_data


def main():
    st.set_page_config(
        page_title="SISCAD - Evaluación Continua",
        page_icon="🧑‍🏫",
        layout="wide",
    )

    st.markdown(
        """
        <style>
        div[data-testid="stAppViewContainer"] {
            background: #f6f8fb;
            color: #0f172a;
        }
        div[data-testid="stSidebar"] {
            background: #ffffff;
            border-right: 1px solid rgba(15, 23, 42, 0.08);
        }
        .css-1d391kg, .css-ffhzg2, .css-1outpf7, .stCard {
            background: #ffffff;
            border: 1px solid rgba(15, 23, 42, 0.08);
            box-shadow: 0 18px 36px rgba(15, 23, 42, 0.06);
            border-radius: 24px;
        }
        .stButton>button {
            background: #111827 !important;
            color: #ffffff !important;
            border: none;
            border-radius: 14px;
            padding: 0.9rem 1.4rem;
            font-weight: 600;
        }
        .stButton>button:hover {
            background: #1e293b !important;
        }
        .stTextInput>div>div>input,
        .stTextArea>div>div>textarea,
        .stSelectbox>div>div>div>div,
        .stNumberInput>div>div>input {
            background: #ffffff;
            color: #0f172a !important;
            border: 1px solid rgba(15, 23, 42, 0.12);
            border-radius: 14px;
        }
        .stTextInput>div>div>input::placeholder,
        .stTextArea>div>div>textarea::placeholder {
            color: rgba(15, 23, 42, 0.38);
        }
        .stMarkdown h1, .stMarkdown h2, .stMarkdown h3, .stMarkdown h4 {
            color: #0f172a !important;
        }
        .stMarkdown p, .stText, .stTextInput>label,
        .stCaption, .stMetric, .css-1aw5q0x, .css-1gkcyyc {
            color: #475569 !important;
        }
        .stSelectbox>div>div>div>div,
        .stMultiSelect>div>div>div,
        .stTextInput>div>div,
        .stTextArea>div>div,
        .stNumberInput>div>div {
            color: #0f172a !important;
        }
        .stSidebar .stButton>button {
            width: 100%;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.title("SISCAD - Evaluación Continua 🧑‍🏫📘")
    st.markdown("## Bienvenido al aula digital del docente")
    st.markdown(
        "Esta app organiza tus evaluaciones con un estilo moderno, claro y profesional. "
        "Usa colores suaves, contraste alto y una interfaz amigable para el trabajo escolar."
    )
    st.markdown("---")

    uploaded_file = st.sidebar.file_uploader("Carga tu archivo Excel (.xlsx)", type=["xlsx"])
    st.sidebar.markdown("### 🗂️ Opciones del profesor")
    st.sidebar.markdown("📌 Usa tu archivo local o sube uno nuevo para gestionar el registro.")
    st.sidebar.markdown("📝 Edita calificaciones, calcula resultados y exporta el formato de secretaría.")
    use_local = False
    if WORKBOOK_FILE.exists():
        use_local = st.sidebar.button("Usar archivo local ListaDeCalificaciones3E.xlsx")

    if uploaded_file is not None:
        try:
            wb = load_workbook(uploaded_file, data_only=True)
            general, students = parse_datos_sheet(wb)
            trimester_data = {sheet: parse_trimester_sheet(wb, sheet) for sheet in TRIMESTER_SHEETS}
            st.session_state["general"] = general
            st.session_state["students"] = students
            st.session_state["trimester_data"] = trimester_data
            st.session_state["source"] = "uploaded"
            st.session_state["template_source"] = uploaded_file.getvalue()
        except Exception as exc:
            st.error(f"No se pudo leer el archivo cargado: {exc}")
    elif use_local:
        try:
            wb = load_workbook(WORKBOOK_FILE, data_only=True)
            general, students = parse_datos_sheet(wb)
            trimester_data = {sheet: parse_trimester_sheet(wb, sheet) for sheet in TRIMESTER_SHEETS}
            st.session_state["general"] = general
            st.session_state["students"] = students
            st.session_state["trimester_data"] = trimester_data
            st.session_state["source"] = "local"
            st.session_state["template_source"] = WORKBOOK_FILE
        except Exception as exc:
            st.error(f"No se pudo leer el archivo local: {exc}")

    if "general" not in st.session_state or "students" not in st.session_state or "trimester_data" not in st.session_state:
        st.warning("Sube el archivo Excel o pulsa el botón para usar el archivo local.")
        st.stop()

    general = st.session_state["general"]
    students = st.session_state["students"]
    trimester_data = st.session_state["trimester_data"]

    st.header("1. Datos generales")
    general["Escuela"] = st.text_input("Escuela", general.get("Escuela", ""))
    general["Docente"] = st.text_input("Docente", general.get("Docente", ""))
    general["Grado"] = st.text_input("Grado", str(general.get("Grado", "")))
    general["Grupo"] = st.text_input("Grupo", general.get("Grupo", ""))
    st.session_state["general"] = general

    st.header("2. Alumnos")
    editor = get_editor()
    if editor is None:
        st.write(students)
    else:
        students = editor(students, num_rows="dynamic")
        st.session_state["students"] = students

    st.header("3. Trimestres")
    selected_trimester = st.selectbox("Selecciona un trimestre", TRIMESTER_SHEETS)
    trim_df = trimester_data[selected_trimester]
    trim_df = normalize_grades(trim_df)

    st.subheader(selected_trimester)
    with st.expander("Actividades"):
        if editor is None:
            st.write(trim_df[["N°", "Nombre"] + ACT_COLUMNS])
        else:
            updated_activities = editor(trim_df[["N°", "Nombre"] + ACT_COLUMNS], num_rows="dynamic")
            for col in ACT_COLUMNS:
                trim_df[col] = updated_activities[col]

    with st.expander("Tareas"):
        if editor is None:
            st.write(trim_df[["N°", "Nombre"] + TASK_COLUMNS])
        else:
            updated_tasks = editor(trim_df[["N°", "Nombre"] + TASK_COLUMNS], num_rows="dynamic")
            for col in TASK_COLUMNS:
                trim_df[col] = updated_tasks[col]

    with st.expander("Proyectos"):
        if editor is None:
            st.write(trim_df[["N°", "Nombre"] + PROJECT_COLUMNS])
        else:
            updated_projects = editor(trim_df[["N°", "Nombre"] + PROJECT_COLUMNS], num_rows="dynamic")
            for col in PROJECT_COLUMNS:
                trim_df[col] = updated_projects[col]

    with st.expander("Participación, Conducta y Examen"):
        if editor is None:
            st.write(trim_df[["N°", "Nombre"] + METRIC_COLUMNS])
        else:
            updated_metrics = editor(trim_df[["N°", "Nombre"] + METRIC_COLUMNS], num_rows="dynamic")
            for col in METRIC_COLUMNS:
                trim_df[col] = updated_metrics[col]

    computed_trim_df = compute_trimester_result(trim_df)
    trimester_data[selected_trimester] = computed_trim_df
    st.session_state["trimester_data"] = trimester_data

    st.subheader("Resultados calculados")
    st.write(computed_trim_df[["N°", "Nombre", "Promedio Actividades", "Promedio Tareas", "Promedio Proyectos", "Participación", "Conducta", "Evaluación (Examen)", "Calif Final"]])

    st.header("4. Resumen general")
    concentrado_df = build_concentrado_sheet(trimester_data)
    st.write(concentrado_df)
    resultados_df = build_resultados_sheet(general, concentrado_df)
    st.write(resultados_df)

    if st.button("Guardar libro actualizado"):
        output_path = Path("registro_evaluacion_actualizado.xlsx")
        template_source = st.session_state.get("template_source", WORKBOOK_FILE)
        save_workbook(general, students, trimester_data, concentrado_df, resultados_df, output_path, template_source=template_source)
        st.success(f"Libro guardado como {output_path.name}")
        with open(output_path, "rb") as f:
            st.download_button(
                label="Descargar archivo actualizado",
                data=f,
                file_name=output_path.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    st.header("5. Exportar formato Secretaría")
    st.write("Selecciona el o los trimestres que necesitas entregar a secretaría.")
    selected_secretaria_trimesters = st.multiselect("Trimestres para exportar", TRIMESTER_SHEETS, default=[selected_trimester])
    if st.button("Generar archivo de Secretaría"):
        if not selected_secretaria_trimesters:
            st.error("Selecciona al menos un trimestre para generar el archivo.")
        else:
            output_path = Path("secretaria_export.xlsx")
            save_secretaria_workbook(selected_secretaria_trimesters, general, trimester_data, output_path)
            st.success(f"Archivo de Secretaría generado: {output_path.name}")
            with open(output_path, "rb") as f:
                st.download_button(
                    label="Descargar archivo de Secretaría",
                    data=f,
                    file_name=output_path.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )


if __name__ == "__main__":
    main()
